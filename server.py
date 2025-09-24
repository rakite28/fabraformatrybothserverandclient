# --- Standard Library Imports ---
import os
import re
import json
import time
import shutil
import sys
import traceback
import uuid
import secrets
import logging
from logging.handlers import RotatingFileHandler
from io import BytesIO
import subprocess
from datetime import datetime, timedelta
from functools import wraps

# --- Third-Party Imports ---
from flask import Flask, jsonify, request, send_from_directory, g, send_file
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import jwt
from PIL import Image
import easyocr
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib import utils, colors
from pydantic import ValidationError, parse_obj_as
from typing import List, Dict

# --- Local Application Imports ---
from database import get_db_connection, init_db
from moderator import NSFWDetector
from validators import (
    LoginModel, RegisterCompanyModel, CreateUserModel, ChangePasswordModel,
    UpdateProfileModel, PrinterModel, ProcessImageModel, GenerateQuotationModel,
    FilamentsPostModel, RefreshTokenModel, SlicerProfileModel, SliceRequestModel
)

# --- Initial Setup ---
try:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    os.chdir(SCRIPT_DIR)
except NameError:
    SCRIPT_DIR = os.getcwd()

CONFIG_PATH = "server_config.json"
APP_CONFIG = {}
ocr_reader = None
nsfw_detector = None

app = Flask(__name__)
CORS(app)

# --- Centralized Logging Setup ---
def setup_logging():
    log_dir = os.path.join(SCRIPT_DIR, 'logs')
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, 'server.log')
    handler = RotatingFileHandler(log_file, maxBytes=1024 * 1024 * 5, backupCount=5)
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s [in %(pathname)s:%(lineno)d]')
    handler.setFormatter(formatter)
    
    root_logger = logging.getLogger()
    if not root_logger.handlers:
        root_logger.addHandler(handler)
        root_logger.setLevel(logging.INFO)
    
    # Configure Flask's logger to use the same handler
    app.logger.handlers = []
    app.logger.addHandler(handler)
    app.logger.setLevel(logging.INFO)
    app.logger.info("Application starting up...")

# --- Global Error Handlers ---
@app.errorhandler(Exception)
def handle_unexpected_error(e):
    app.logger.error(f"An unhandled exception occurred: {e}", exc_info=True)
    return jsonify({
        "error": "An unexpected server error occurred.",
        "message": "The issue has been logged for investigation."
    }), 500

@app.errorhandler(ValidationError)
def handle_validation_error(e):
    app.logger.warning(f"Validation error for request on '{request.path}': {e.errors()}")
    error_details = {err['loc'][0]: err['msg'] for err in e.errors()}
    return jsonify({
        "error": "Input validation failed",
        "message": "One or more fields are invalid.",
        "details": error_details
    }), 400

# --- Flask App Context & Database ---
@app.before_request
def before_request():
    g.db = get_db_connection()

@app.teardown_request
def teardown_request(exception):
    db = g.pop('db', None)
    if db is not None:
        db.close()

# --- Helper Functions ---
def get_ocr_reader():
    global ocr_reader
    if ocr_reader is None:
        app.logger.info("⏳ Loading EasyOCR model into memory...")
        try:
            ocr_reader = easyocr.Reader(['en'], gpu=True)
            app.logger.info("✅ EasyOCR model loaded.")
        except Exception as e:
            app.logger.critical(f"🛑 FATAL: Could not load EasyOCR model. Error: {e}")
            ocr_reader = None
    return ocr_reader

def get_nsfw_detector():
    global nsfw_detector
    if nsfw_detector is None:
        app.logger.info("⏳ Loading NSFW Detector model...")
        nsfw_detector = NSFWDetector()
        app.logger.info("✅ NSFW Detector loaded.")
    return nsfw_detector

def get_company_data_path(company_id, *args):
    base_path = os.path.join(SCRIPT_DIR, "data", str(company_id))
    os.makedirs(base_path, exist_ok=True)
    return os.path.join(base_path, *args)

def load_app_config():
    try:
        with open(CONFIG_PATH, 'r') as f: return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError): return {}

def save_app_config(data):
    with open(CONFIG_PATH, 'w') as f: json.dump(data, f, indent=4)
    global APP_CONFIG
    APP_CONFIG = data.copy()

def get_safe_path(subpath):
    share_dir = os.path.abspath(APP_CONFIG.get("SERVER_SHARE_DIR", "server_share"))
    target_path = os.path.abspath(os.path.join(share_dir, subpath))
    if not target_path.startswith(share_dir): return None
    return target_path

# --- Slicer Helper Functions ---
def find_profile_path(company_id, profile_type, filename):
    safe_filename = secure_filename(filename)
    # Check user-specific (private) profiles first
    user_profile_dir = get_company_data_path(company_id, "slicer_profiles", profile_type)
    user_profile_path = os.path.join(user_profile_dir, safe_filename)
    if os.path.exists(user_profile_path):
        return user_profile_path

    # Fallback to system-wide (public) profiles
    system_profile_base = APP_CONFIG.get("SLICER_SYSTEM_PROFILE_PATH")
    if not system_profile_base or not os.path.exists(system_profile_base):
        return None

    for root, _, files in os.walk(system_profile_base):
        if safe_filename in files:
            return os.path.join(root, safe_filename)
            
    return None

def parse_gcode_for_data(gcode_path):
    filament_grams = 0.0
    print_time_hours = 0.0
    with open(gcode_path, 'r', errors='ignore') as f: content = f.read()

    filament_match = re.search(r';\s*filament used \[g\]\s*=\s*([\d\.]+)', content)
    if filament_match: filament_grams = float(filament_match.group(1))
    
    time_match = re.search(r';\s*estimated printing time .*\s*=\s*(?:(\d+)h\s*)?(?:(\d+)m\s*)?(?:(\d+)s)?', content)
    if time_match:
        h = int(time_match.group(1) or 0)
        m = int(time_match.group(2) or 0)
        s = int(time_match.group(3) or 0)
        print_time_hours = h + (m / 60.0) + (s / 3600.0)
        
    return {"filament_grams": round(filament_grams, 2), "print_time_hours": round(print_time_hours, 4)}

def run_slicer(stl_path, machine_profile, filament_profile, process_profile):
    slicer_exe = APP_CONFIG.get("SLICER_EXECUTABLE_PATH")
    if not slicer_exe or not os.path.exists(slicer_exe):
        return {"success": False, "error": "Slicer executable not configured or not found."}
    
    output_dir = os.path.dirname(stl_path)
    
    command = [
        slicer_exe, "--slice", stl_path,
        "--load", machine_profile,
        "--load", filament_profile,
        "--load", process_profile,
        "--output", output_dir
    ]
    try:
        app.logger.info(f"Running slicer command: {' '.join(command)}")
        # Using CREATE_NO_WINDOW for Windows to prevent console pop-up
        creation_flags = subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0
        result = subprocess.run(command, capture_output=True, text=True, check=True, creationflags=creation_flags)
        
        gcode_files = [f for f in os.listdir(output_dir) if f.endswith('.gcode')]
        if not gcode_files:
            app.logger.error(f"Slicing stdout: {result.stdout}\nSlicing stderr: {result.stderr}")
            return {"success": False, "error": "Slicing completed, but no G-code file was generated."}
            
        gcode_path = os.path.join(output_dir, gcode_files[0])
        parsed_data = parse_gcode_for_data(gcode_path)
        
        # Clean up generated files
        for f in os.listdir(output_dir):
            if f.endswith(('.gcode', '.3mf')):
                os.remove(os.path.join(output_dir, f))

        return {"success": True, "data": parsed_data}
        
    except subprocess.CalledProcessError as e:
        error_message = f"OrcaSlicer failed with exit code {e.returncode}:\n{e.stderr}"
        app.logger.error(error_message)
        user_friendly_error = "Slicing failed. The STL file may be invalid, or the object may be too large for the selected printer's build plate."
        return {"success": False, "error": user_friendly_error}
    except Exception as e:
        app.logger.error(f"An unexpected error occurred during slicing: {e}", exc_info=True)
        return {"success": False, "error": "An unexpected server error occurred during slicing."}

# --- Decorators ---
def validate_with(model: any):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            try:
                # Handle different content types
                if request.is_json:
                    json_data = request.get_json()
                elif request.form:
                    json_data = json.loads(request.form.get('json', '{}'))
                else:
                    return jsonify({"error": "Unsupported Media Type. Expected JSON or multipart/form-data."}), 415

                # Pydantic can parse lists directly
                if isinstance(json_data, list):
                    g.validated_data = parse_obj_as(model, json_data)
                else:
                    g.validated_data = model.parse_obj(json_data)
                
                return f(*args, **kwargs)
            except ValidationError as e:
                raise e  # Let the global error handler catch this
            except (json.JSONDecodeError, TypeError):
                return jsonify({"error": "Invalid JSON in request body or form data."}), 400
            except Exception as e:
                app.logger.error(f"Error parsing request: {e}", exc_info=True)
                return jsonify({"error": "Invalid request format."}), 400
        return decorated_function
    return decorator

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        if 'Authorization' in request.headers:
            token = request.headers['Authorization'].split(" ")[1]
        if not token:
            return jsonify({'message': 'Token is missing!'}), 401
        try:
            data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=["HS256"])
            g.current_user = {
                'user_id': data['user_id'], 'username': data['username'],
                'company_id': data['company_id'], 'role': data['role']
            }
        except jwt.ExpiredSignatureError:
            return jsonify({'message': 'Token has expired!'}), 401
        except Exception as e:
            app.logger.warning(f"Invalid token received: {e}")
            return jsonify({'message': 'Token is invalid!'}), 401
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    @token_required
    def decorated(*args, **kwargs):
        if g.current_user['role'] != 'admin':
            return jsonify({'message': 'Admin privileges required!'}), 403
        return f(*args, **kwargs)
    return decorated

# --- Authentication & Registration Endpoints ---
@app.route('/auth/companies', methods=['GET'])
def get_companies():
    companies_cur = g.db.execute("SELECT id, name FROM companies ORDER BY name").fetchall()
    return jsonify([dict(row) for row in companies_cur])

@app.route('/auth/login', methods=['POST'])
@validate_with(LoginModel)
def login():
    data = g.validated_data
    user_row = g.db.execute(
        "SELECT * FROM users WHERE lower(email) = lower(?) OR lower(username) = lower(?)", 
        (data.identifier, data.identifier)
    ).fetchone()

    if not user_row:
        return jsonify({'message': 'User not found'}), 401
    
    user = dict(user_row)
    if check_password_hash(user['password_hash'], data.password):
        access_token = jwt.encode({
            'user_id': user['id'], 'username': user['username'], 'company_id': user['company_id'],
            'role': user['role'], 'exp': datetime.utcnow() + timedelta(hours=24)
        }, app.config['SECRET_KEY'], algorithm="HS256")

        response_data = {'token': access_token}

        if data.remember_me:
            remember_token = secrets.token_hex(32)
            token_hash = generate_password_hash(remember_token)
            expires_at = datetime.utcnow() + timedelta(days=30)
            
            try:
                g.db.execute(
                    "INSERT INTO auth_tokens (user_id, token_hash, expires_at) VALUES (?, ?, ?)",
                    (user['id'], token_hash, expires_at)
                )
                g.db.commit()
                response_data['remember_token'] = remember_token
            except Exception as e:
                g.db.rollback()
                app.logger.error(f"Could not save remember token for user {user['id']}: {e}")
        
        return jsonify(response_data)
    
    return jsonify({'message': 'Invalid credentials'}), 401

@app.route('/auth/refresh', methods=['POST'])
@validate_with(RefreshTokenModel)
def refresh():
    data = g.validated_data
    all_tokens = g.db.execute("SELECT user_id, token_hash, expires_at FROM auth_tokens").fetchall()
    user_id = None
    
    for row in all_tokens:
        if check_password_hash(row['token_hash'], data.remember_token):
            if datetime.utcnow() < datetime.fromisoformat(row['expires_at'].replace(' ', 'T')):
                user_id = row['user_id']
                break
            else: # Token is expired, delete it
                g.db.execute("DELETE FROM auth_tokens WHERE token_hash = ?", (row['token_hash'],))
                g.db.commit()
                return jsonify({'message': 'Remember token has expired'}), 401

    if not user_id:
        return jsonify({'message': 'Invalid or expired remember token'}), 401

    user = g.db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        return jsonify({'message': 'User associated with token not found'}), 404

    access_token = jwt.encode({
        'user_id': user['id'], 'username': user['username'], 'company_id': user['company_id'],
        'role': user['role'], 'exp': datetime.utcnow() + timedelta(hours=24)
    }, app.config['SECRET_KEY'], algorithm="HS256")
    
    return jsonify({'token': access_token})

@app.route('/auth/logout', methods=['POST'])
@token_required
def logout():
    data = request.json
    remember_token = data.get('remember_token')
    if remember_token:
        all_tokens = g.db.execute("SELECT token_hash FROM auth_tokens WHERE user_id = ?", (g.current_user['user_id'],)).fetchall()
        for row in all_tokens:
            if check_password_hash(row['token_hash'], remember_token):
                g.db.execute("DELETE FROM auth_tokens WHERE token_hash = ?", (row['token_hash'],))
                g.db.commit()
                break
    return jsonify({'message': 'Logout successful'}), 200

@app.route('/auth/register_company', methods=['POST'])
@validate_with(RegisterCompanyModel)
def register_company():
    data = g.validated_data
    if g.db.execute("SELECT id FROM companies WHERE lower(name) = lower(?)", (data.company_name,)).fetchone():
        return jsonify({'message': 'A company with this name already exists'}), 409
    if g.db.execute("SELECT id FROM users WHERE lower(email) = lower(?)", (data.admin_email,)).fetchone():
        return jsonify({'message': 'This email is already registered'}), 409

    try:
        new_company_id = str(uuid.uuid4())
        password_hash = generate_password_hash(data.admin_password)
        cursor = g.db.cursor()
        cursor.execute("INSERT INTO companies (id, name) VALUES (?, ?)", (new_company_id, data.company_name))
        cursor.execute("INSERT INTO users (id, username, email, password_hash, company_id, role) VALUES (?, ?, ?, ?, ?, ?)",
                         (str(uuid.uuid4()), data.admin_username, data.admin_email, password_hash, new_company_id, 'admin'))
        g.db.commit()
    except Exception as e:
        g.db.rollback()
        app.logger.error(f"Company registration failed: {e}")
        return jsonify({'message': 'An error occurred during registration.'}), 500
    
    return jsonify({'message': f"Company '{data.company_name}' created successfully."}), 201

@app.route('/auth/create_user', methods=['POST'])
@admin_required
@validate_with(CreateUserModel)
def create_user():
    data = g.validated_data
    company_id = g.current_user['company_id']
    if g.db.execute("SELECT id FROM users WHERE lower(email) = lower(?)", (data.email,)).fetchone():
        return jsonify({'message': 'This email is already registered'}), 409
    if g.db.execute("SELECT id FROM users WHERE lower(username) = lower(?) AND company_id = ?", (data.username, company_id)).fetchone():
        return jsonify({'message': 'This username already exists in your company'}), 409

    try:
        password_hash = generate_password_hash(data.password)
        g.db.execute("INSERT INTO users (id, username, email, password_hash, company_id, role) VALUES (?, ?, ?, ?, ?, ?)",
                       (str(uuid.uuid4()), data.username, data.email, password_hash, company_id, data.role))
        g.db.commit()
    except Exception as e:
        g.db.rollback()
        app.logger.error(f"User creation failed for company {company_id}: {e}")
        return jsonify({'message': 'Failed to create user.'}), 500

    return jsonify({'message': f"User '{data.username}' created successfully."}), 201

# --- User Profile Endpoints ---
@app.route('/user/profile', methods=['GET', 'POST'])
@token_required
def user_profile():
    user_id = g.current_user['user_id']
    if request.method == 'GET':
        user_row = g.db.execute(
            "SELECT username, email, phone_number, dob, profile_picture_path FROM users WHERE id = ?", (user_id,)
        ).fetchone()
        if not user_row:
            return jsonify({'message': 'User not found'}), 404
        
        profile_data = dict(user_row)
        if profile_data.get('profile_picture_path'):
            profile_data['profile_picture_url'] = f"{request.url_root.rstrip('/')}/user/profile_picture/{profile_data['profile_picture_path']}"
        return jsonify(profile_data)

    if request.method == 'POST':
        @validate_with(UpdateProfileModel)
        def handle_post():
            data = g.validated_data
            update_fields = data.dict(exclude_unset=True)
            
            if not update_fields:
                return jsonify({'message': 'No update information provided'}), 400

            set_clause = ", ".join([f"{key} = ?" for key in update_fields.keys()])
            params = list(update_fields.values()) + [user_id]

            try:
                g.db.execute(f"UPDATE users SET {set_clause} WHERE id = ?", tuple(params))
                g.db.commit()
                return jsonify({'message': 'Profile updated successfully'})
            except Exception as e:
                g.db.rollback()
                app.logger.error(f"Profile update failed for user {user_id}: {e}")
                return jsonify({'message': 'Failed to update profile.'}), 500
        return handle_post()

@app.route('/user/change_password', methods=['POST'])
@token_required
@validate_with(ChangePasswordModel)
def change_password():
    data = g.validated_data
    user_id = g.current_user['user_id']
    
    user = g.db.execute("SELECT password_hash FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        return jsonify({'message': 'User not found'}), 404
    
    if not check_password_hash(user['password_hash'], data.current_password):
        return jsonify({'message': 'Current password is not correct'}), 403
        
    new_password_hash = generate_password_hash(data.new_password)
    try:
        g.db.execute("UPDATE users SET password_hash = ? WHERE id = ?", (new_password_hash, user_id))
        g.db.commit()
        return jsonify({'message': 'Password updated successfully'})
    except Exception as e:
        g.db.rollback()
        app.logger.error(f"Password change failed for user {user_id}: {e}")
        return jsonify({'message': 'Failed to update password.'}), 500

@app.route('/user/profile_picture', methods=['POST'])
@token_required
def upload_profile_picture():
    user_id = g.current_user['user_id']
    company_id = g.current_user['company_id']

    if 'file' not in request.files or not request.files['file'].filename:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']

    # --- NSFW Check ---
    detector = get_nsfw_detector()
    file_bytes = file.read()
    file.seek(0) # Reset file pointer after reading
    is_safe, reason = detector.is_image_safe(file_bytes)
    if not is_safe:
        app.logger.warning(f"User {user_id} tried to upload NSFW profile picture. Reason: {reason}")
        return jsonify({'error': 'Image was flagged as inappropriate and was rejected.'}), 400

    filename = secure_filename(f"{user_id}_{int(time.time())}{os.path.splitext(file.filename)[1]}")
    upload_folder = get_company_data_path(company_id, "profile_pictures")
    file_path = os.path.join(upload_folder, filename)
    file.save(file_path)

    try:
        g.db.execute("UPDATE users SET profile_picture_path = ? WHERE id = ?", (filename, user_id))
        g.db.commit()
        new_url = f"{request.url_root.rstrip('/')}/user/profile_picture/{filename}"
        return jsonify({'message': 'Profile picture updated', 'filepath': filename, 'url': new_url})
    except Exception as e:
        g.db.rollback()
        app.logger.error(f"Failed to update profile picture path in DB for user {user_id}: {e}")
        return jsonify({'message': 'Failed to update database.'}), 500

@app.route('/user/profile_picture/<path:filename>')
@token_required
def serve_profile_picture(filename):
    company_id = g.current_user['company_id']
    directory = get_company_data_path(company_id, "profile_pictures")
    return send_from_directory(directory, filename)

# --- Core Application Endpoints ---
@app.route('/server/settings', methods=['GET', 'POST'])
@admin_required
def handle_server_settings():
    if request.method == 'POST':
        try:
            save_app_config(request.json)
            return jsonify({"status": "success", "message": "Settings saved."})
        except Exception as e:
            app.logger.error(f"Failed to save server settings: {e}")
            return jsonify({"status": "error", "message": "Failed to save settings."}), 500
    else:
        return jsonify(load_app_config())

@app.route('/server/files/', defaults={'subpath': ''})
@app.route('/server/files/<path:subpath>')
@admin_required
def list_files(subpath):
    safe_path = get_safe_path(subpath)
    if not safe_path or not os.path.isdir(safe_path): return jsonify({"error": "Invalid path"}), 404
    try:
        file_list = [{"name": item, "type": "dir" if os.path.isdir(os.path.join(safe_path, item)) else "file",
                      "size": os.path.getsize(os.path.join(safe_path, item)) if not os.path.isdir(os.path.join(safe_path, item)) else 0}
                     for item in os.listdir(safe_path)]
        return jsonify(file_list)
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route('/server/upload/', defaults={'subpath': ''}, methods=['POST'])
@app.route('/server/upload/<path:subpath>', methods=['POST'])
@admin_required
def upload_file(subpath):
    safe_path = get_safe_path(subpath)
    if not safe_path or not os.path.isdir(safe_path): return jsonify({"error": "Invalid destination"}), 400
    if 'file' not in request.files or not request.files['file'].filename: return jsonify({"error": "No file part"}), 400
    file = request.files['file']; filename = secure_filename(file.filename)
    file.save(os.path.join(safe_path, filename))
    return jsonify({"status": "success", "message": f"File '{filename}' uploaded."})

@app.route('/server/download/<path:filepath>')
@admin_required
def download_server_file(filepath):
    safe_path = get_safe_path(filepath)
    if not safe_path or not os.path.isfile(safe_path): return jsonify({"error": "File not found"}), 404
    return send_from_directory(os.path.dirname(safe_path), os.path.basename(safe_path), as_attachment=True)

@app.route('/printers', methods=['GET', 'POST'])
@token_required
def handle_printers():
    company_id = g.current_user['company_id']
    if request.method == 'POST':
        @validate_with(List[PrinterModel])
        def handle_post():
            printers_data = g.validated_data
            try:
                cursor = g.db.cursor()
                cursor.execute("DELETE FROM printers WHERE company_id = ?", (company_id,))
                for p in printers_data:
                    cursor.execute("""
                        INSERT INTO printers (id, company_id, brand, model, setup_cost, maintenance_cost, lifetime_years, power_w, price_kwh, buffer_factor, uptime_percent)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (p.id, company_id, p.brand, p.model, p.setup_cost, p.maintenance_cost, p.lifetime_years, p.power_w, p.price_kwh, p.buffer_factor, p.uptime_percent))
                g.db.commit()
                return jsonify({"status": "saved"})
            except Exception as e:
                g.db.rollback()
                app.logger.error(f"Failed to save printers for company {company_id}: {e}")
                return jsonify({"status": "error", "message": "Failed to save printers."}), 500
        return handle_post()
    else: # GET
        printers_cur = g.db.execute("SELECT * FROM printers WHERE company_id = ?", (company_id,)).fetchall()
        return jsonify([dict(row) for row in printers_cur])

@app.route('/filaments', methods=['GET', 'POST'])
@token_required
def handle_filaments():
    company_id = g.current_user['company_id']
    if request.method == 'POST':
        @validate_with(Dict[str, Dict[str, FilamentsPostModel]])
        def handle_post():
            data = g.validated_data
            try:
                cursor = g.db.cursor()
                cursor.execute("DELETE FROM filaments WHERE company_id = ?", (company_id,))
                for material, brands in data.items():
                    for brand, details in brands.items():
                        cursor.execute("INSERT INTO filaments (company_id, material, brand, price, stock_g, efficiency_factor) VALUES (?, ?, ?, ?, ?, ?)",
                                       (company_id, material, brand, details.price, details.stock_g, details.efficiency_factor))
                g.db.commit()
                return jsonify({"status": "saved"})
            except Exception as e:
                g.db.rollback()
                app.logger.error(f"Failed to save filaments for company {company_id}: {e}")
                return jsonify({"status": "error", "message": "Failed to save filaments."}), 500
        return handle_post()
    else: # GET
        filaments_cur = g.db.execute("SELECT * FROM filaments WHERE company_id = ?", (company_id,)).fetchall()
        filaments_dict = {}
        for row in filaments_cur:
            material = row['material']
            if material not in filaments_dict: filaments_dict[material] = {}
            filaments_dict[material][row['brand']] = {'price': row['price'], 'stock_g': row['stock_g'], 'efficiency_factor': row['efficiency_factor']}
        return jsonify(filaments_dict)

@app.route('/logs', methods=['GET'])
@token_required
def get_logs():
    log_path = get_company_data_path(g.current_user['company_id'], "app_logs.json")
    try:
        with open(log_path, 'r') as f: return jsonify(json.load(f))
    except (FileNotFoundError, json.JSONDecodeError): return jsonify([])

@app.route('/processed_log', methods=['GET'])
@token_required
def get_processed_log():
    log_path = get_company_data_path(g.current_user['company_id'], "processed_log.json")
    try:
        with open(log_path, 'r') as f: return jsonify(json.load(f))
    except (FileNotFoundError, json.JSONDecodeError): return jsonify({})

@app.route('/images/<path:filename>')
@token_required
def serve_image(filename):
    image_dir = get_company_data_path(g.current_user['company_id'], "local_log_images")
    return send_from_directory(image_dir, filename)

@app.route('/download/log/<path:filename>')
@token_required
def download_log_file(filename):
    excel_dir = get_company_data_path(g.current_user['company_id'], "Excel_Logs")
    return send_from_directory(os.path.abspath(excel_dir), filename, as_attachment=True)

@app.route('/download/masterlog/<year_month>')
@token_required
def download_master_log_file(year_month):
    filename = f"master_log_{year_month.split('_')[1]}.xlsx"
    directory = get_company_data_path(g.current_user['company_id'], "Monthly_Expenditure", year_month)
    return send_from_directory(os.path.abspath(directory), filename, as_attachment=True)
    
# --- Processing and Quotation Endpoints ---
@app.route('/quotation/slice_model', methods=['POST'])
@token_required
def slice_model():
    company_id = g.current_user['company_id']
    if 'stl_file' not in request.files:
        return jsonify({"error": "Missing 'stl_file' in request"}), 400
    
    try:
        profiles = SliceRequestModel.parse_obj(request.form)
    except ValidationError as e:
        raise e

    stl_file = request.files['stl_file']
    if not stl_file.filename.lower().endswith('.stl'):
        return jsonify({"error": "Invalid file type. Only .stl is supported."}), 400
    
    temp_dir = get_company_data_path(company_id, "temp_slice", str(uuid.uuid4()))
    os.makedirs(temp_dir, exist_ok=True)
    
    safe_stl_filename = secure_filename(stl_file.filename)
    stl_path = os.path.join(temp_dir, safe_stl_filename)
    stl_file.save(stl_path)

    # Find full paths for profile files
    machine_profile_path = find_profile_path(company_id, "machine", profiles.machine_profile)
    filament_profile_path = find_profile_path(company_id, "filament", profiles.filament_profile)
    process_profile_path = find_profile_path(company_id, "process", profiles.process_profile)
    
    if not all([machine_profile_path, filament_profile_path, process_profile_path]):
        shutil.rmtree(temp_dir)
        missing = [
            p[0] for p in [
                ("machine", machine_profile_path), ("filament", filament_profile_path), ("process", process_profile_path)
            ] if not p[1]
        ]
        return jsonify({"error": f"Could not find the following profiles: {', '.join(missing)}"}), 404

    slicer_result = run_slicer(stl_path, machine_profile_path, filament_profile_path, process_profile_path)
    
    # Clean up temporary directory
    shutil.rmtree(temp_dir)

    if slicer_result["success"]:
        return jsonify(slicer_result["data"])
    else:
        return jsonify({"error": slicer_result["error"]}), 500

@app.route('/generate_quotation', methods=['POST'])
@token_required
@validate_with(GenerateQuotationModel)
def generate_quotation():
    company_id = g.current_user['company_id']
    data = g.validated_data.dict() # Convert Pydantic model to dict
    
    # PDF generation needs a logo path if one was uploaded.
    # This endpoint assumes the logo is already stored and its path is provided.
    if data.get("company_details", {}).get("logo_path"):
        # For security, ensure path is within company data.
        logo_filename = os.path.basename(data["company_details"]["logo_path"])
        safe_logo_path = get_company_data_path(company_id, "uploads", logo_filename)
        if os.path.exists(safe_logo_path):
            data["company_details"]["logo_path"] = safe_logo_path
        else:
            data["company_details"]["logo_path"] = None

    buffer = BytesIO()
    generate_quotation_pdf(buffer, data)
    buffer.seek(0)

    customer_name_safe = re.sub(r'[^a-zA-Z0-9_]', '', data['customer_name'].replace(' ', '_'))
    filename = f"Quotation_{customer_name_safe}_{int(time.time())}.pdf"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )

@app.route('/ocr_upload', methods=['POST'])
@token_required
def ocr_upload():
    if 'image' not in request.files: return jsonify({"error": "No image file provided"}), 400
    
    image_file = request.files['image']
    image_bytes = image_file.read()
    image_file.seek(0)
    
    # --- NSFW Check ---
    detector = get_nsfw_detector()
    is_safe, reason = detector.is_image_safe(image_bytes)
    if not is_safe:
        app.logger.warning(f"User {g.current_user['user_id']} tried to upload NSFW image for OCR. Reason: {reason}")
        return jsonify({'error': 'Image was flagged as inappropriate and was rejected.'}), 400
    
    reader = get_ocr_reader()
    if not reader: return jsonify({"error": "OCR model not available."}), 500
    
    ocr_results = reader.readtext(image_bytes)
    return jsonify(extract_data_from_ocr(g.current_user['company_id'], ocr_results))

@app.route('/process_image', methods=['POST'])
@token_required
@validate_with(ProcessImageModel)
def process_image_upload():
    company_id = g.current_user['company_id']
    final_data = g.validated_data.dict(by_alias=True)

    if 'image' not in request.files:
        return jsonify({"error": "Missing image file"}), 400
    image_file = request.files['image']
    
    # --- NSFW Check ---
    image_bytes = image_file.read()
    image_file.seek(0)
    detector = get_nsfw_detector()
    is_safe, reason = detector.is_image_safe(image_bytes)
    if not is_safe:
        app.logger.warning(f"User {g.current_user['user_id']} tried to process NSFW image. Reason: {reason}")
        return jsonify({'error': 'Image was flagged as inappropriate and was rejected.'}), 400

    # Step 1: Save the uploaded image
    image_dir = get_company_data_path(company_id, "local_log_images")
    new_filename = final_data["Filename"] + os.path.splitext(image_file.filename)[1]
    image_file.save(os.path.join(image_dir, new_filename))

    # Step 2: Validate Printer and Filament data from the database
    printer_row = g.db.execute("SELECT * FROM printers WHERE id=? AND company_id=?", (final_data.get("printer_id"), company_id)).fetchone()
    filament_row = g.db.execute("SELECT * FROM filaments WHERE material=? AND brand=? AND company_id=?", (final_data.get("Material"), final_data.get("Brand"), company_id)).fetchone()
    
    if not printer_row or not filament_row:
        return jsonify({"status": "error", "message": "Critical data missing: Printer or filament not found in the database."}), 400
    
    printer, filament = dict(printer_row), dict(filament_row)

    # Step 3: Perform Calculations
    cogs = calculate_cogs_values(final_data, printer, filament)

    # Step 4: Create Individual Excel Log
    excel_path, excel_msg = create_excel_file(company_id, final_data, printer, filament)
    if not excel_path:
        return jsonify({"status": "error", "message": f"Failed to create Excel log: {excel_msg}"}), 500

    # Step 5: Update Master Log
    success, msg = log_to_master_excel(company_id, excel_path, final_data, cogs['user_cogs'], cogs['default_cogs'])
    if not success:
        return jsonify({"status": "error", "message": f"Failed to update master log: {msg}"}), 500
        
    # Step 6: Finalize and commit changes
    update_filament_stock(company_id, final_data)
    save_app_log(company_id, final_data, cogs, new_filename)
    
    processed_log_path = get_company_data_path(company_id, "processed_log.json")
    processed_log = json.load(open(processed_log_path)) if os.path.exists(processed_log_path) else {}
    processed_log[os.path.basename(image_file.filename)] = "completed"
    with open(processed_log_path, 'w') as f: json.dump(processed_log, f, indent=2)
    
    return jsonify({"status": "success", "message": "File processed and logged successfully."})
    
# --- Processing Helper Functions (with logging) ---
def parse_time_string(time_str):
    h_match = re.search(r'(\d+)\s*h', time_str, re.IGNORECASE); h = int(h_match.group(1)) if h_match else 0
    m_match = re.search(r'(\d+)\s*m', time_str, re.IGNORECASE); m = int(m_match.group(1)) if m_match else 0
    s_match = re.search(r'(\d+)\s*s', time_str, re.IGNORECASE); s = int(s_match.group(1)) if s_match else 0
    return round(h + (m / 60.0) + (s / 3600.0), 4)

def calculate_printer_hourly_rate(printer_data):
    try:
        total_cost = printer_data['setup_cost'] + (printer_data['maintenance_cost'] * printer_data['lifetime_years'])
        total_hours = printer_data['lifetime_years'] * 365 * 24 * (printer_data.get('uptime_percent', 50) / 100)
        if total_hours == 0: return 0.0
        return (total_cost / total_hours) + ((printer_data['power_w'] / 1000) * printer_data['price_kwh'])
    except (KeyError, TypeError, ZeroDivisionError) as e:
        app.logger.warning(f"Could not calculate printer hourly rate: {e}")
        return 0.0

def calculate_cogs_values(form_data, printer_data, filament_data):
    try:
        filament_g = float(form_data.get("Filament (g)", 0)); time_str = form_data.get("Time (e.g. 7h 30m)", "0h 0m")
        labour_time_min = float(form_data.get("Labour Time (min)", 0)); labour_rate_user = float(form_data.get("Labour Rate (₹/hr)", 0))
        print_time_hours = parse_time_string(time_str)
        mat_cost = (filament_data.get('price', 0) / 1000) * filament_g * filament_data.get('efficiency_factor', 1.0)
        labour_cogs = (labour_rate_user / 60) * labour_time_min
        printer_cogs = calculate_printer_hourly_rate(printer_data) * printer_data.get('buffer_factor', 1.0) * print_time_hours
        total_cogs_user = mat_cost + labour_cogs + printer_cogs
        mat_cost_default = (filament_data.get('price', 0) / 1000) * filament_g
        labour_cogs_default = (100 / 60) * labour_time_min
        printer_cogs_default = calculate_printer_hourly_rate(printer_data) * print_time_hours
        total_cogs_default = mat_cost_default + labour_cogs_default + printer_cogs_default
        return {"user_cogs": total_cogs_user, "default_cogs": total_cogs_default}
    except (ValueError, TypeError, KeyError, ZeroDivisionError) as e:
        app.logger.warning(f"Could not calculate COGS values: {e}")
        return {"user_cogs": 0.0, "default_cogs": 0.0}

def extract_data_from_ocr(company_id, ocr_results):
    full_text = " ".join([item[1] for item in ocr_results]).lower()
    extracted_data = {"filament": 0.0, "time_str": "0h 0m", "material": None, "detected_printer_id": None}
    filament_g = 0.0
    priority_match = re.search(r'total filament\D*(\d+\.?\d*)\s*g', full_text)
    if priority_match: filament_g = round(float(priority_match.group(1)), 2)
    else:
        all_g_matches = re.findall(r'(\d+\.?\d*)\s*g', full_text)
        if all_g_matches:
            try:
                numeric_values = [float(val) for val in all_g_matches]
                if numeric_values: filament_g = round(max(numeric_values), 2)
            except (ValueError, TypeError): filament_g = 0.0
    extracted_data["filament"] = filament_g
    
    hours, minutes = 0, 0
    time_block_match = re.search(r'(?:total time|print time)\D*(?:(\d+)\s*h)?\s*(?:(\d+)\s*m)?', full_text)
    if time_block_match:
        h_val, m_val = time_block_match.groups()
        if h_val: hours = int(h_val)
        if m_val: minutes = int(m_val)
    if hours == 0 and minutes == 0:
        h_values = [int(h) for h in re.findall(r'(\d+)\s*h', full_text)]
        m_values = [int(m) for m in re.findall(r'(\d+)\s*m', full_text)]
        if h_values: hours = max(h_values)
        if m_values: minutes = max(m_values)
    if hours > 0 or minutes > 0: extracted_data["time_str"] = f"{hours}h {minutes}m"

    filaments_cur = g.db.execute("SELECT DISTINCT material FROM filaments WHERE company_id = ?", (company_id,)).fetchall()
    known_materials = [row['material'].lower() for row in filaments_cur]
    for material in known_materials:
        if re.search(r'\b' + re.escape(material) + r'\b', full_text):
            extracted_data["material"] = material.upper(); break

    printers_cur = g.db.execute("SELECT id, brand, model FROM printers WHERE company_id = ?", (company_id,)).fetchall()
    for printer in printers_cur:
        if printer['brand'].lower() in full_text or printer['model'].lower() in full_text:
            extracted_data["detected_printer_id"] = printer['id']; break
    return extracted_data

def update_filament_stock(company_id, final_data):
    try:
        material, brand = final_data.get("Material"), final_data.get("Brand")
        grams_used = float(final_data.get("Filament (g)", 0))
        if not all([material, brand, grams_used > 0]): return
        g.db.execute("UPDATE filaments SET stock_g = stock_g - ? WHERE company_id = ? AND material = ? AND brand = ?",
                       (grams_used, company_id, material, brand))
        g.db.commit()
    except Exception as e:
        g.db.rollback()
        app.logger.error(f"Error updating stock for company {company_id}: {e}")

def create_excel_file(company_id, final_data, printer, filament):
    try:
        template_path = APP_CONFIG.get("TEMPLATE_PATH", "FDM.xlsx")
        if not os.path.exists(template_path): return None, f"Template '{template_path}' not found."
        wb = load_workbook(template_path)
        excel_output_dir = get_company_data_path(company_id, "Excel_Logs"); os.makedirs(excel_output_dir, exist_ok=True)
        new_path = os.path.join(excel_output_dir, f"{final_data['Filename']}.xlsx")
        calc_ws, adv_ws = wb["Calculation Sheet"], wb["Adv. Inputs"]
        calc_ws['D4'] = final_data["Filename"]; calc_ws['D6'] = datetime.fromisoformat(final_data["timestamp"])
        calc_ws['D7'] = "FabraForma"; calc_ws['D9'] = final_data["Material"]
        calc_ws['D10'] = float(final_data["Filament Cost (₹/kg)"]); calc_ws['D11'] = float(final_data["Filament (g)"])
        calc_ws['D12'] = parse_time_string(final_data["Time (e.g. 7h 30m)"]); calc_ws['D13'] = float(final_data["Labour Time (min)"])
        adv_ws['C6'] = float(final_data.get("Labour Rate (₹/hr)", 100)); adv_ws['D6'] = 100
        adv_ws['C11'] = printer['setup_cost']; adv_ws['D11'] = printer['setup_cost']
        adv_ws['C15'] = printer['maintenance_cost']; adv_ws['D15'] = printer['maintenance_cost']
        adv_ws['C18'] = printer['lifetime_years']; adv_ws['D18'] = printer['lifetime_years']
        adv_ws['C22'] = printer['power_w']; adv_ws['D22'] = printer['power_w']
        adv_ws['C23'] = printer['price_kwh']; adv_ws['D23'] = printer['price_kwh']
        adv_ws['C4'] = filament.get('efficiency_factor', 1.0); adv_ws['D4'] = 1.0
        adv_ws['C28'] = printer.get('buffer_factor', 1.0); adv_ws['D28'] = 1.0
        wb.save(new_path)
        return new_path, "Success"
    except Exception as e:
        app.logger.error(f"Error in create_excel_file: {e}", exc_info=True)
        return None, "An unexpected error occurred while creating the Excel file."

def log_to_master_excel(company_id, file_path, final_data, user_cogs, default_cogs):
    try:
        source_wb = load_workbook(file_path, data_only=True); calc_ws = source_wb["Calculation Sheet"]
        date_val = calc_ws["D6"].value or datetime.now()
        config = load_app_config()
        values = [calc_ws[cell].value for cell in config["cells"]]
        p_num = os.path.splitext(os.path.basename(file_path))[0]
        new_row = [None, date_val, p_num] + values + [user_cogs, default_cogs, f'=HYPERLINK("{os.path.abspath(file_path)}", "Source File")']
        month_name = date_val.strftime("%B")
        ym_folder = get_company_data_path(company_id, "Monthly_Expenditure", f"{date_val.year}_{month_name}")
        os.makedirs(ym_folder, exist_ok=True)
        master_path = os.path.join(ym_folder, f"master_log_{month_name}.xlsx")
        if os.path.exists(master_path):
            master_wb = load_workbook(master_path); master_ws = master_wb.active
            for row_idx in range(master_ws.max_row, 1, -1):
                if master_ws.cell(row=row_idx, column=2).value == "TOTALS": master_ws.delete_rows(row_idx); break
            all_rows = [list(row) for row in master_ws.iter_rows(min_row=2, values_only=True) if row and row[2] != p_num]
        else:
            master_wb = Workbook(); master_ws = master_wb.active; master_ws.title = "DataLog"
            master_ws.append(config["headers"]); all_rows = []
        all_rows.append(new_row)
        all_rows.sort(key=lambda row: (row[1] if isinstance(row[1], datetime) else datetime.min, str(row[2])))
        if master_ws.max_row > 1: master_ws.delete_rows(2, master_ws.max_row)
        for idx, row_data in enumerate(all_rows, start=1):
            row_data[0] = idx
            if isinstance(row_data[1], datetime): row_data[1] = row_data[1].strftime("%Y-%m-%d %H:%M:%S")
            master_ws.append(row_data)
        last_row = master_ws.max_row; totals_row_idx = last_row + 1
        master_ws.cell(row=totals_row_idx, column=2, value="TOTALS").font = Font(bold=True)
        cols_to_sum = ["Filament (g)", "Time (h)", "Labour Time (min)", "User COGS (₹)", "Default COGS (₹)"]
        header_row = [cell.value for cell in master_ws[1]]
        for col_name in cols_to_sum:
            try:
                col_idx = header_row.index(col_name) + 1
                formula = f"=SUM({get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{last_row})"
                master_ws.cell(row=totals_row_idx, column=col_idx, value=formula).font = Font(bold=True)
            except ValueError: app.logger.warning(f"Master log missing header '{col_name}'.")
        master_wb.save(master_path)
        return True, f"Logged to master: {os.path.basename(file_path)}"
    except Exception as e:
        app.logger.error(f"Error in log_to_master_excel: {e}", exc_info=True)
        return False, "An unexpected error occurred while logging to the master file."

def save_app_log(company_id, final_data, cogs_data, local_image_filename):
    log_path = get_company_data_path(company_id, "app_logs.json")
    try:
        logs = []
        if os.path.exists(log_path):
            with open(log_path, 'r') as f: content = f.read(); logs = json.loads(content) if content else []
        log_entry = {
            "timestamp": final_data["timestamp"], "filename": final_data["Filename"], "image_path": local_image_filename,
            "data": { "Printer": final_data["Printer"], "Material": final_data["Material"], "Brand": final_data["Brand"],
                      "Filament (g)": final_data["Filament (g)"], "Time": final_data["Time (e.g. 7h 30m)"],
                      "User COGS (₹)": f"{cogs_data['user_cogs']:.2f}", "Default COGS (₹)": f"{cogs_data['default_cogs']:.2f}" }}
        logs.append(log_entry); logs.sort(key=lambda x: x['timestamp'], reverse=True)
        with open(log_path, 'w') as f: json.dump(logs, f, indent=4)
    except Exception as e: app.logger.error(f"FAILED to save app log for company {company_id}: {e}")

def generate_quotation_pdf(buffer, data):
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    comp_details = data.get("company_details", {})
    if comp_details.get("logo_path") and os.path.exists(comp_details["logo_path"]):
        try:
            img = utils.ImageReader(comp_details["logo_path"])
            i_width, i_height = img.getSize()
            aspect = i_height / float(i_width)
            c.drawImage(img, 40, height - 100, width=80, height=(80 * aspect))
        except Exception as e:
            app.logger.error(f"Could not draw logo on PDF: {e}")
    c.setFont("Helvetica-Bold", 16); c.drawRightString(width - 50, height - 60, comp_details.get("name", "Your Company"))
    c.setFont("Helvetica", 10); c.drawRightString(width - 50, height - 75, comp_details.get("address", "Company Address"))
    c.drawRightString(width - 50, height - 90, comp_details.get("contact", "Contact Info"))
    c.setFont("Helvetica-Bold", 24); c.drawString(50, height - 150, "Quotation")
    c.line(50, height - 155, width - 50, height - 155)
    c.setFont("Helvetica-Bold", 12); c.drawString(50, height - 190, "BILLED TO:")
    c.setFont("Helvetica", 12); c.drawString(50, height - 205, data.get("customer_name", "Valued Customer"))
    if data.get("customer_company"): c.drawString(50, height - 220, data.get("customer_company"))
    c.setFont("Helvetica", 12); c.drawRightString(width - 50, height - 190, f"Date: {datetime.now().strftime('%Y-%m-%d')}")
    y_position = height - 260; c.setFont("Helvetica-Bold", 11)
    c.drawString(60, y_position, "Part Description"); c.drawRightString(width - 200, y_position, "Unit Price"); c.drawRightString(width - 50, y_position, "Total")
    c.line(50, y_position - 10, width - 50, y_position - 10); y_position -= 30
    total_cogs = sum(part.get("cogs", 0) for part in data["parts"])
    margin_percent = data.get("margin_percent", 0)
    subtotal = total_cogs / (1 - (margin_percent / 100.0)) if margin_percent < 100 else 0
    tax_rate_percent = data.get("tax_rate_percent", 0); tax_amount = subtotal * (tax_rate_percent / 100.0)
    grand_total = subtotal + tax_amount
    c.setFont("Helvetica", 10)
    line_item_description = f"{len(data['parts'])} Custom Manufactured Part(s)"
    c.drawString(60, y_position, line_item_description); c.drawRightString(width - 200, y_position, f"Rs{subtotal:,.2f}"); c.drawRightString(width - 50, y_position, f"Rs{subtotal:,.2f}")
    y_position -= 30; c.line(width - 250, y_position, width - 50, y_position); y_position -= 20
    c.setFont("Helvetica", 11); c.drawRightString(width - 200, y_position, "Subtotal:"); c.drawRightString(width - 50, y_position, f"Rs{subtotal:,.2f}")
    y_position -= 20; c.drawRightString(width - 200, y_position, f"Tax ({tax_rate_percent}%):"); c.drawRightString(width - 50, y_position, f"Rs{tax_amount:,.2f}")
    y_position -= 20; c.setFont("Helvetica-Bold", 12); c.drawRightString(width - 200, y_position, "Grand Total:"); c.drawRightString(width - 50, y_position, f"Rs{grand_total:,.2f}")
    c.setFont("Helvetica-Oblique", 9); c.drawString(50, 50, "Thank you for your business! Prices are valid for 30 days.")
    c.showPage(); c.save()

# --- App Initialization ---
def initialize_app():
    global APP_CONFIG
    setup_logging()
    APP_CONFIG = load_app_config()
    defaults = {
        "SERVER_SHARE_DIR": "server_share",
        "TEMPLATE_PATH": "FDM.xlsx",
        "SLICER_EXECUTABLE_PATH": r"C:\Program Files\OrcaSlicer\orca-slicer.exe",
        "SLICER_SYSTEM_PROFILE_PATH": r"C:\Users\YourUser\AppData\Roaming\OrcaSlicer\user",
        "cells": ["D4", "D9", "D10", "D11", "D12", "D13"],
        "headers": ["Sr. No", "Date", "Part Number", "Filename", "Material", "Filament Cost (₹/kg)", "Filament (g)", "Time (h)", "Labour Time (min)", "User COGS (₹)", "Default COGS (₹)", "Source Link"]
    }
    app.config['SECRET_KEY'] = APP_CONFIG.get('SECRET_KEY', secrets.token_hex(32))
    
    config_updated = False
    for key, value in defaults.items():
        if key not in APP_CONFIG:
            APP_CONFIG[key] = value
            config_updated = True
            
    if config_updated or 'SECRET_KEY' not in APP_CONFIG:
        APP_CONFIG['SECRET_KEY'] = app.config['SECRET_KEY']
        save_app_config(APP_CONFIG)
    
    os.makedirs(os.path.join(SCRIPT_DIR, "data"), exist_ok=True)
    os.makedirs(APP_CONFIG["SERVER_SHARE_DIR"], exist_ok=True)
    
    with app.app_context():
        init_db(SCRIPT_DIR)

# --- Run Application ---
initialize_app()
# Pre-load models on startup
get_ocr_reader()
get_nsfw_detector()

if __name__ == "__main__":
    from waitress import serve
    app.logger.info("--- Starting Production Server with Waitress ---")
    app.run(host='0.0.0.0', port=5000, debug =True, use_reloader=False)