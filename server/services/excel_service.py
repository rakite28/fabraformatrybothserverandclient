import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from flask import current_app

from ..helpers import get_company_data_path
from .calculation_service import parse_time_string

def create_excel_file(company_id, final_data, printer, filament, config):
    try:
        template_path = config.get("TEMPLATE_PATH")
        if not template_path or not os.path.exists(template_path):
            error_msg = f"Excel template not found at path: {template_path}"
            current_app.logger.error(error_msg)
            return None, error_msg

        wb = load_workbook(template_path)
        excel_output_dir = get_company_data_path(company_id, "Excel_Logs")
        os.makedirs(excel_output_dir, exist_ok=True)
        new_path = os.path.join(excel_output_dir, f"{final_data['Filename']}.xlsx")

        calc_ws, adv_ws = wb["Calculation Sheet"], wb["Adv. Inputs"]
        calc_ws['D4'] = final_data["Filename"]
        calc_ws['D6'] = datetime.fromisoformat(final_data["timestamp"])
        calc_ws['D7'] = "FabraForma"
        calc_ws['D9'] = final_data["Material"]
        calc_ws['D10'] = float(final_data["Filament Cost (₹/kg)"])
        calc_ws['D11'] = float(final_data["Filament (g)"])
        calc_ws['D12'] = parse_time_string(final_data["Time (e.g. 7h 30m)"])
        calc_ws['D13'] = float(final_data["Labour Time (min)"])

        adv_ws['C6'] = float(final_data.get("Labour Rate (₹/hr)", 100))
        adv_ws['D6'] = 100
        adv_ws['C11'] = printer['setup_cost']
        adv_ws['D11'] = printer['setup_cost']
        adv_ws['C15'] = printer['maintenance_cost']
        adv_ws['D15'] = printer['maintenance_cost']
        adv_ws['C18'] = printer['lifetime_years']
        adv_ws['D18'] = printer['lifetime_years']
        adv_ws['C22'] = printer['power_w']
        adv_ws['D22'] = printer['power_w']
        adv_ws['C23'] = printer['price_kwh']
        adv_ws['D23'] = printer['price_kwh']
        adv_ws['C4'] = filament.get('efficiency_factor', 1.0)
        adv_ws['D4'] = 1.0
        adv_ws['C28'] = printer.get('buffer_factor', 1.0)
        adv_ws['D28'] = 1.0

        wb.save(new_path)
        return new_path, "Success"
    except Exception as e:
        current_app.logger.error(f"Error in create_excel_file: {e}", exc_info=True)
        return None, "An unexpected error occurred while creating the Excel file."

def log_to_master_excel(company_id, file_path, final_data, user_cogs, default_cogs, config):
    try:
        source_wb = load_workbook(file_path, data_only=True)
        calc_ws = source_wb["Calculation Sheet"]
        date_val = calc_ws["D6"].value or datetime.now()

        values = [calc_ws[cell].value for cell in config["cells"]]
        p_num = os.path.splitext(os.path.basename(file_path))[0]
        new_row = [None, date_val, p_num] + values + [user_cogs, default_cogs, f'=HYPERLINK("{os.path.abspath(file_path)}", "Source File")']

        month_name = date_val.strftime("%B")
        ym_folder = get_company_data_path(company_id, "Monthly_Expenditure", f"{date_val.year}_{month_name}")
        os.makedirs(ym_folder, exist_ok=True)
        master_path = os.path.join(ym_folder, f"master_log_{month_name}.xlsx")

        if os.path.exists(master_path):
            master_wb = load_workbook(master_path)
            master_ws = master_wb.active
            for row_idx in range(master_ws.max_row, 1, -1):
                if master_ws.cell(row=row_idx, column=2).value == "TOTALS":
                    master_ws.delete_rows(row_idx)
                    break
            all_rows = [list(row) for row in master_ws.iter_rows(min_row=2, values_only=True) if row and row[2] != p_num]
        else:
            master_wb = Workbook()
            master_ws = master_wb.active
            master_ws.title = "DataLog"
            master_ws.append(config["headers"])
            all_rows = []

        all_rows.append(new_row)
        all_rows.sort(key=lambda row: (row[1] if isinstance(row[1], datetime) else datetime.min, str(row[2])))

        if master_ws.max_row > 1:
            master_ws.delete_rows(2, master_ws.max_row)

        for idx, row_data in enumerate(all_rows, start=1):
            row_data[0] = idx
            if isinstance(row_data[1], datetime):
                row_data[1] = row_data[1].strftime("%Y-%m-%d %H:%M:%S")
            master_ws.append(row_data)

        last_row = master_ws.max_row
        totals_row_idx = last_row + 1
        master_ws.cell(row=totals_row_idx, column=2, value="TOTALS").font = Font(bold=True)

        cols_to_sum = ["Filament (g)", "Time (h)", "Labour Time (min)", "User COGS (₹)", "Default COGS (₹)"]
        header_row = [cell.value for cell in master_ws[1]]

        for col_name in cols_to_sum:
            try:
                col_idx = header_row.index(col_name) + 1
                formula = f"=SUM({get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{last_row})"
                master_ws.cell(row=totals_row_idx, column=col_idx, value=formula).font = Font(bold=True)
            except ValueError:
                current_app.logger.warning(f"Master log missing header '{col_name}'.")

        master_wb.save(master_path)
        return True, f"Logged to master: {os.path.basename(file_path)}"
    except Exception as e:
        current_app.logger.error(f"Error in log_to_master_excel: {e}", exc_info=True)
        return False, "An unexpected error occurred while logging to the master file."